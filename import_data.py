"""
import_data.py
-------------
Standalone script to initialise the database and import reimbursement data
from two Excel workbooks:
  - "Theresa Reimbursement_To 19 May 2025.xlsx"
  - "Tony Claim_To 16 Dec 2025.xlsx"

Run from the project root:
    python import_data.py
"""

import sys
import os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app import app, db
from models import Company, User, ClaimPeriod, TravelClaim, PurchaseClaim
import bcrypt
import openpyxl
from datetime import datetime, date

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

THERESA_FILE = os.path.join(BASE_DIR, "Theresa Reimbursement_To 19 May 2025.xlsx")
THERESA_SHEETS = ["To 18 Sep 2025", "To 18 May 2025 Paid"]

TONY_FILE = os.path.join(BASE_DIR, "Tony Claim_To 16 Dec 2025.xlsx")
TONY_SHEETS = [
    "July-Paid",
    "August-Paid",
    "September-Paid",
    "Nov",
    "Oct",
    "Nov 24 - Apr 2025-Paid",
    "May 2025-23 June 2025-Paid",
    "24 June 2025-13 Oct 2025_Paid",
    "14 Oct 2025-18 Dec 2025",
]

MILEAGE_RATE = 0.88  # $/km

# Summary sentinel strings that indicate end of data
TRAVEL_STOP_SENTINELS = {"sub total", "subtotal"}
PURCHASE_STOP_SENTINELS = {
    "business account",
    "personal account",
    "total",
    "total purchases",
    "claim total",
    "business",
    "personal",
}


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------


def create_password_hash(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def parse_date(value) -> date | None:
    """Convert various date representations to a Python date object."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        if not value or value.startswith("="):
            return None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                pass
        # Last-resort: let dateutil try (if available)
        try:
            from dateutil import parser as dp
            return dp.parse(value, dayfirst=True).date()
        except Exception:
            pass
    return None


def parse_number(value, default=0.0) -> float | None:
    """Convert a cell value to float; returns None for formula strings."""
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        s = value.strip()
        if s.startswith("="):
            return None  # caller must handle formula
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return default
    return default


def is_stop_sentinel(value, sentinels: set) -> bool:
    """Return True if the cell value matches a known stop sentinel."""
    if value is None:
        return False
    return str(value).strip().lower() in sentinels


def cell_str(value) -> str | None:
    """Return stripped string or None."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def load_sheet_rows(workbook_path: str, sheet_name: str):
    """Load all rows from a sheet as a list of tuples (values_only=True)."""
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"  [WARN] Sheet '{sheet_name}' not found in {os.path.basename(workbook_path)}")
        return []
    ws = wb[sheet_name]
    return list(ws.iter_rows(values_only=True))


# ---------------------------------------------------------------------------
# Theresa parsing
# ---------------------------------------------------------------------------

def parse_theresa_travel(rows, claim_period: ClaimPeriod) -> int:
    """
    Parse Part 1 travel claims from Theresa's sheet.

    Header at row index 7 (row 8 in Excel):
      Col B=1 Date | C=2 From | D=3 To | E=4 Purpose | F=5 Trip Type
      G=6 Toll | H=7 Distance (km) | I=8 Mileage Cost

    Data starts at row index 8.
    Stop when col B is "Sub total" or both col A and col B are None.
    """
    count = 0
    data_start = 8  # 0-indexed

    for i, row in enumerate(rows):
        if i < data_start:
            continue

        # Safely extend row tuple if short
        row = list(row) + [None] * max(0, 12 - len(row))

        col_b = row[1]  # date
        col_a = row[0]

        # Stop conditions
        if is_stop_sentinel(col_b, TRAVEL_STOP_SENTINELS):
            break
        if col_a is None and col_b is None:
            break

        try:
            parsed_date = parse_date(col_b)
            if parsed_date is None:
                continue  # skip rows with no valid date

            origin = cell_str(row[2])
            destination = cell_str(row[3])
            purpose = cell_str(row[4])
            trip_type = cell_str(row[5])
            toll_cost = parse_number(row[6], default=0.0)
            distance = parse_number(row[7], default=0.0)

            # Mileage cost: may be a formula string
            raw_mileage = row[8]
            if isinstance(raw_mileage, str) and raw_mileage.strip().startswith("="):
                # Calculate from distance
                mileage_cost = (distance or 0.0) * MILEAGE_RATE
            else:
                mileage_cost = parse_number(raw_mileage, default=None)
                if mileage_cost is None:
                    mileage_cost = (distance or 0.0) * MILEAGE_RATE

            tc = TravelClaim(
                claim_period_id=claim_period.id,
                date=parsed_date,
                origin=origin,
                destination=destination,
                purpose=purpose,
                trip_type=trip_type,
                toll_cost=toll_cost or 0.0,
                distance_km=distance or 0.0,
                mileage_cost=round(mileage_cost, 4),
            )
            db.session.add(tc)
            count += 1

        except Exception as exc:
            print(f"    [SKIP] Travel row {i + 1}: {exc}")
            continue

    return count


def parse_theresa_purchases(rows, claim_period: ClaimPeriod) -> int:
    """
    Parse Part 2 purchases from Theresa's sheet.

    Header around row index 29 (row 30 in Excel):
      Col B=1 Date | C=2 Purchased From | D=3 Items | E=4 Purpose
      F=5 Price | G=6 Receipt

    We locate the header by searching for it from row 20 onwards.
    Data starts immediately after the header row.
    Stop at summary rows or end of data.
    """
    count = 0
    header_idx = None

    # Search for the purchase section header (look for "Date" in col B around rows 20-40)
    for i, row in enumerate(rows):
        if i < 20:
            continue
        if i > 50:
            break
        row_ext = list(row) + [None] * max(0, 10 - len(row))
        b = cell_str(row_ext[1])
        if b and b.lower() == "date":
            # Verify this is the purchases header by checking nearby cols
            header_idx = i
            break

    if header_idx is None:
        # Fall back: start at row index 29
        header_idx = 29

    data_start = header_idx + 1

    for i, row in enumerate(rows):
        if i < data_start:
            continue

        row = list(row) + [None] * max(0, 10 - len(row))

        col_b = row[1]
        col_c = row[2]
        col_d = row[3]
        col_e = row[4]
        col_f = row[5]
        col_g = row[6]

        # Stop at summary sentinels
        if is_stop_sentinel(col_b, PURCHASE_STOP_SENTINELS):
            break
        if is_stop_sentinel(col_c, PURCHASE_STOP_SENTINELS):
            break
        if is_stop_sentinel(col_d, PURCHASE_STOP_SENTINELS):
            break

        # Skip entirely empty rows
        if all(v is None for v in [col_b, col_c, col_d, col_e, col_f]):
            continue

        try:
            parsed_date = parse_date(col_b)
            purchased_from = cell_str(col_c)
            item = cell_str(col_d)
            purpose = cell_str(col_e)
            price = parse_number(col_f, default=0.0)
            receipt_val = cell_str(col_g)
            has_receipt = receipt_val is not None and receipt_val.strip().lower() in ("y", "yes")

            # If all meaningful fields are empty, skip
            if parsed_date is None and item is None and purchased_from is None and (price or 0) == 0:
                continue

            pc = PurchaseClaim(
                claim_period_id=claim_period.id,
                date=parsed_date,
                item=item,
                purchased_from=purchased_from,
                purpose=purpose,
                price=price or 0.0,
                has_receipt=has_receipt,
                category="Business",
            )
            db.session.add(pc)
            count += 1

        except Exception as exc:
            print(f"    [SKIP] Purchase row {i + 1}: {exc}")
            continue

    return count


def import_theresa(user: User, company: Company) -> None:
    """Import all sheets from Theresa's workbook."""
    print(f"\nImporting Theresa's data from: {os.path.basename(THERESA_FILE)}")

    for sheet_name in THERESA_SHEETS:
        status = "paid" if "paid" in sheet_name.lower() else "pending"
        print(f"  Sheet: '{sheet_name}' → status={status}")

        rows = load_sheet_rows(THERESA_FILE, sheet_name)
        if not rows:
            print(f"  [WARN] No rows loaded for sheet '{sheet_name}', skipping.")
            continue

        cp = ClaimPeriod(
            user_id=user.id,
            company_id=company.id,
            period_name=sheet_name,
            status=status,
        )
        db.session.add(cp)
        db.session.flush()  # get cp.id

        travel_count = parse_theresa_travel(rows, cp)
        purchase_count = parse_theresa_purchases(rows, cp)

        print(f"    Travel claims: {travel_count}, Purchase claims: {purchase_count}")


# ---------------------------------------------------------------------------
# Tony parsing
# ---------------------------------------------------------------------------

def parse_tony_travel(rows, claim_period: ClaimPeriod) -> int:
    """
    Parse Part 1 travel claims from Tony's sheet.

    Header at row index 7 (row 8 in Excel):
      Col B=1 Date | C=2 Origin | D=3 Destination | E=4 Purpose
      F=5 Trip Type | G=6 Toll | H=7 Distance (km)

    Mileage cost = distance * 0.88 (no mileage column in Tony's sheet).
    Data starts at row index 8.
    Stop at "Sub total" or empty date.
    """
    count = 0
    data_start = 8  # 0-indexed

    for i, row in enumerate(rows):
        if i < data_start:
            continue

        row = list(row) + [None] * max(0, 12 - len(row))

        col_b = row[1]  # date
        col_a = row[0]

        # Stop conditions
        if is_stop_sentinel(col_b, TRAVEL_STOP_SENTINELS):
            break
        if col_a is None and col_b is None:
            break

        try:
            parsed_date = parse_date(col_b)
            if parsed_date is None:
                continue

            origin = cell_str(row[2])
            destination = cell_str(row[3])
            purpose = cell_str(row[4])
            trip_type = cell_str(row[5])
            toll_cost = parse_number(row[6], default=0.0)
            distance = parse_number(row[7], default=0.0)
            mileage_cost = round((distance or 0.0) * MILEAGE_RATE, 4)

            tc = TravelClaim(
                claim_period_id=claim_period.id,
                date=parsed_date,
                origin=origin,
                destination=destination,
                purpose=purpose,
                trip_type=trip_type,
                toll_cost=toll_cost or 0.0,
                distance_km=distance or 0.0,
                mileage_cost=mileage_cost,
            )
            db.session.add(tc)
            count += 1

        except Exception as exc:
            print(f"    [SKIP] Travel row {i + 1}: {exc}")
            continue

    return count


def parse_tony_purchases(rows, claim_period: ClaimPeriod) -> int:
    """
    Parse Part 2 purchases from Tony's sheet.

    Header at row index 17 (row 18 in Excel):
      Col B=1 Date | C=2 Item(s) | D=3 Purchased From | E=4 Purpose
      F=5 Price | G=6 Receipt

    Stop at summary rows.
    """
    count = 0
    header_idx = None

    # Search for the purchase section header from row index 14 onward
    for i, row in enumerate(rows):
        if i < 14:
            continue
        if i > 30:
            break
        row_ext = list(row) + [None] * max(0, 10 - len(row))
        b = cell_str(row_ext[1])
        if b and b.lower() == "date":
            header_idx = i
            break

    if header_idx is None:
        header_idx = 17  # fall back to row 18 (0-indexed 17)

    data_start = header_idx + 1

    for i, row in enumerate(rows):
        if i < data_start:
            continue

        row = list(row) + [None] * max(0, 10 - len(row))

        col_b = row[1]
        col_c = row[2]
        col_d = row[3]
        col_e = row[4]
        col_f = row[5]
        col_g = row[6]

        # Stop at summary sentinels
        if is_stop_sentinel(col_b, PURCHASE_STOP_SENTINELS):
            break
        if is_stop_sentinel(col_c, PURCHASE_STOP_SENTINELS):
            break
        if is_stop_sentinel(col_d, PURCHASE_STOP_SENTINELS):
            break

        # Skip entirely empty rows
        if all(v is None for v in [col_b, col_c, col_d, col_e, col_f]):
            continue

        try:
            parsed_date = parse_date(col_b)
            item = cell_str(col_c)
            purchased_from = cell_str(col_d)
            purpose = cell_str(col_e)
            price = parse_number(col_f, default=0.0)
            receipt_val = cell_str(col_g)
            has_receipt = receipt_val is not None and receipt_val.strip().lower() in ("y", "yes")

            # If all meaningful fields are empty, skip
            if parsed_date is None and item is None and purchased_from is None and (price or 0) == 0:
                continue

            pc = PurchaseClaim(
                claim_period_id=claim_period.id,
                date=parsed_date,
                item=item,
                purchased_from=purchased_from,
                purpose=purpose,
                price=price or 0.0,
                has_receipt=has_receipt,
                category="Business",
            )
            db.session.add(pc)
            count += 1

        except Exception as exc:
            print(f"    [SKIP] Purchase row {i + 1}: {exc}")
            continue

    return count


def import_tony(user: User, company: Company) -> None:
    """Import all sheets from Tony's workbook."""
    print(f"\nImporting Tony's data from: {os.path.basename(TONY_FILE)}")

    # Load the workbook once to get actual sheet names (handles casing / whitespace)
    try:
        wb = openpyxl.load_workbook(TONY_FILE, data_only=True)
        actual_sheets = wb.sheetnames
        wb.close()
    except Exception as exc:
        print(f"  [ERROR] Could not open Tony's file: {exc}")
        return

    for sheet_name in TONY_SHEETS:
        # Match against actual sheet names (case-insensitive fallback)
        if sheet_name not in actual_sheets:
            matched = next(
                (s for s in actual_sheets if s.strip().lower() == sheet_name.strip().lower()),
                None,
            )
            if matched is None:
                print(f"  [WARN] Sheet '{sheet_name}' not found, skipping.")
                continue
            sheet_name = matched  # use exact name from workbook

        status = "paid" if "paid" in sheet_name.lower() else "pending"
        print(f"  Sheet: '{sheet_name}' → status={status}")

        rows = load_sheet_rows(TONY_FILE, sheet_name)
        if not rows:
            print(f"  [WARN] No rows loaded for sheet '{sheet_name}', skipping.")
            continue

        cp = ClaimPeriod(
            user_id=user.id,
            company_id=company.id,
            period_name=sheet_name,
            status=status,
        )
        db.session.add(cp)
        db.session.flush()  # get cp.id

        travel_count = parse_tony_travel(rows, cp)
        purchase_count = parse_tony_purchases(rows, cp)

        print(f"    Travel claims: {travel_count}, Purchase claims: {purchase_count}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    with app.app_context():
        print("Dropping and recreating all database tables...")
        db.drop_all()
        db.create_all()
        print("Tables created.")

        # -- Company ----------------------------------------------------------
        company = Company(name="Tivoli International (Aust) Pty Ltd")
        db.session.add(company)
        db.session.flush()
        print(f"Created company: {company.name}")

        # -- Users ------------------------------------------------------------
        admin_user = User(
            username="admin",
            role="admin",
            full_name="Administrator",
            company_id=company.id,
            password_hash=create_password_hash("admin123"),
        )
        theresa_user = User(
            username="theresa",
            role="user",
            full_name="Theresa Xu",
            company_id=company.id,
            password_hash=create_password_hash("password123"),
        )
        tony_user = User(
            username="tony",
            role="user",
            full_name="Tony",
            company_id=company.id,
            password_hash=create_password_hash("password123"),
        )
        db.session.add_all([admin_user, theresa_user, tony_user])
        db.session.flush()
        print("Created users: admin, theresa, tony")

        # -- Import data ------------------------------------------------------
        import_theresa(theresa_user, company)
        import_tony(tony_user, company)

        # -- Commit -----------------------------------------------------------
        db.session.commit()
        print("\nDatabase initialised successfully!")
        print("Users created: admin/admin123, theresa/password123, tony/password123")
